import openpyxl
import re
import json
import urllib.request
import urllib.error
from datetime import datetime

EXCEL_PATH = r"c:\Users\raghu\Desktop\ksaw\public\MCE NEW DATA.xlsx"
SUPABASE_URL = "https://wgtzcjsajncrvibtlhxv.supabase.co"
SUPABASE_KEY = "sb_publishable_ipyAcrU1KLwKoS20uWRBdA_iMVuIWtx"

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation"
}

def clean_str(val):
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() == "none" or s.lower() == "null" else s

def clean_aadhaar(val):
    s = clean_str(val)
    return re.sub(r"\D", "", s)

def clean_phone(val):
    s = clean_str(val)
    s = s.replace(".0", "")
    digits = re.sub(r"\D", "", s)
    if len(digits) > 10:
        digits = digits[-10:]
    return digits

def parse_dob(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = clean_str(val)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None

def extract_pin(address):
    if not address:
        return None
    m = re.search(r"\b(\d{6})\b", address)
    return m.group(1) if m else None

def split_name(full_name):
    clean = clean_str(full_name)
    if not clean:
        return "Unknown", "."
    parts = clean.split()
    if len(parts) == 1:
        return parts[0], "."
    return parts[0], " ".join(parts[1:])

def get_current_max_ref():
    url = f"{SUPABASE_URL}/rest/v1/registrations?select=reference_number"
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req) as resp:
        rows = json.loads(resp.read().decode())
    max_num = 0
    for r in rows:
        ref = r.get("reference_number") or ""
        digits = re.findall(r"\d+", ref)
        if digits:
            max_num = max(max_num, int(digits[0]))
    return max_num

def fetch_existing_aadhaars():
    url = f"{SUPABASE_URL}/rest/v1/registrations?select=aadhaar_number,reference_number"
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req) as resp:
        rows = json.loads(resp.read().decode())
    existing = {}
    for r in rows:
        num = clean_aadhaar(r.get("aadhaar_number"))
        if num:
            existing[num] = r.get("reference_number")
    return existing

def run():
    print("--- Fetching current database status ---")
    current_max_ref = get_current_max_ref()
    existing_db = fetch_existing_aadhaars()
    next_ref_num = max(401, current_max_ref + 1)
    print(f"Current Max Reference in DB: {current_max_ref}")
    print(f"Next Reference ID will start at: KSAW {next_ref_num:03d}")

    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheet = wb.active

    headers = [clean_str(sheet.cell(1, c).value) for c in range(1, sheet.max_column + 1)]
    col_map = {name: idx + 1 for idx, name in enumerate(headers) if name}

    records_to_insert = []
    seen_in_file = set()
    skipped_db_dup = 0
    skipped_file_dup = 0

    for r in range(2, sheet.max_row + 1):
        def get_val(header_name):
            col_idx = col_map.get(header_name)
            if col_idx:
                return sheet.cell(r, col_idx).value
            return None

        # Student Name
        full_name = get_val("STUDENT FULL NAME")
        if not full_name:
            continue
        first_name, last_name = split_name(full_name)

        # Aadhaar
        aadhaar = clean_aadhaar(get_val("Aadhaar CARD NUMBER"))
        if aadhaar:
            if aadhaar in existing_db:
                skipped_db_dup += 1
                continue
            if aadhaar in seen_in_file:
                skipped_file_dup += 1
                continue
            seen_in_file.add(aadhaar)

        # Father Name
        father_name = get_val("FATHER NAME")
        g_first, g_last = split_name(father_name)

        # Phone
        phone = clean_phone(get_val("MOBILE NUMBER"))

        # DOB & Gender
        dob = parse_dob(get_val("DATE OF BIRTH"))
        gender = clean_str(get_val("GENDER")).capitalize()
        if gender not in ("Male", "Female", "Other"):
            gender = "Male" if "m" in gender.lower() else "Female"

        # Address & PIN
        full_address = clean_str(get_val("FULL ADDRESS WITH PIN CODE"))
        pin = extract_pin(full_address)
        district = clean_str(get_val("DISTRICT"))
        taluk = clean_str(get_val("TALUK"))

        # Education & Stream
        qualification = clean_str(get_val("QUALIFICATION"))
        branch = clean_str(get_val("BRANCH"))
        course = clean_str(get_val("PREFERED COURSE NAME"))

        # Caste details
        rd_number = clean_str(get_val("RD NUMBER OF CASTE CERTIFICATE"))
        caste = clean_str(get_val("CASTE")).upper()
        sub_caste = clean_str(get_val("SUB CASTE")).upper()
        nigama = clean_str(get_val("NIGAMA")).upper()

        # URLs
        sslc_url = clean_str(get_val("SSLC MARKSCARD"))
        puc_url = clean_str(get_val("PUC OR Diploma MARKSCARD"))
        aadhaar_url = clean_str(get_val("Aadhaar CARD"))
        caste_url = clean_str(get_val("CASTE CERTIFICATE"))
        photo_url = clean_str(get_val("PASSPORT SIZE PHOTO"))

        # Proof mapping rules:
        # age_proof = Aadhaar Card link
        age_proof = aadhaar_url if aadhaar_url else None
        # education_proof = Highest qualification (PUC / Diploma if present, else SSLC)
        education_proof = puc_url if puc_url else (sslc_url if sslc_url else None)

        ref_id = f"KSAW {next_ref_num:03d}"
        next_ref_num += 1

        email = f"{phone or 'student'}@ksaw.org"

        row_payload = {
            "reference_number": ref_id,
            "saf_number": None,
            "first_name": first_name,
            "last_name": last_name,
            "email": email,
            "phone": phone or "0000000000",
            "dob": dob,
            "gender": gender,
            "institution_name": "MALNAD COLLEGE OF ENGINEERING HASSAN",
            "center_location": "HASSAN",
            "guardianship": "Father",
            "guardian_salutation": "Mr",
            "guardian_first_name": g_first,
            "guardian_last_name": g_last,
            "cur_street1": full_address or district or "Karnataka",
            "cur_district": district or None,
            "cur_taluk": taluk or None,
            "cur_state": "Karnataka",
            "cur_zip": pin,
            "per_street1": full_address or district or "Karnataka",
            "per_district": district or None,
            "per_taluk": taluk or None,
            "per_state": "Karnataka",
            "per_zip": pin,
            "same_address": "Yes",
            "education": qualification or None,
            "stream": branch or None,
            "skill_sought": course or None,
            "aadhaar_number": aadhaar or None,
            "rd_number": rd_number or None,
            "caste": caste or None,
            "caste_sub_category": sub_caste or None,
            "nigama": nigama or None,
            "age_proof": age_proof,
            "education_proof": education_proof,
            "aadhaar_proof": aadhaar_url or None,
            "caste_proof": caste_url or None,
            "profile_image": photo_url or None,
            "declaration_accepted": True,
            "status": "Pending"
        }
        records_to_insert.append(row_payload)

    print(f"Total valid records prepared: {len(records_to_insert)}")
    print(f"Skipped DB duplicates: {skipped_db_dup}")
    print(f"Skipped intra-file duplicates: {skipped_file_dup}")
    if records_to_insert:
        print(f"Reference ID Range: {records_to_insert[0]['reference_number']} to {records_to_insert[-1]['reference_number']}")

    # Batch insert into Supabase
    BATCH_SIZE = 20
    inserted_count = 0
    for i in range(0, len(records_to_insert), BATCH_SIZE):
        batch = records_to_insert[i:i + BATCH_SIZE]
        req_body = json.dumps(batch).encode("utf-8")
        insert_url = f"{SUPABASE_URL}/rest/v1/registrations"
        req = urllib.request.Request(insert_url, data=req_body, headers=HEADERS, method="POST")
        try:
            with urllib.request.urlopen(req) as resp:
                res_data = json.loads(resp.read().decode())
                inserted_count += len(res_data)
                print(f"Inserted batch {i//BATCH_SIZE + 1}: {len(res_data)} records (Total: {inserted_count})")
        except urllib.error.HTTPError as e:
            err_msg = e.read().decode()
            print(f"HTTP Error {e.code} inserting batch {i//BATCH_SIZE + 1}: {err_msg}")
            raise e

    print(f"\nSUCCESS! Successfully dumped {inserted_count} records from MCE NEW DATA.xlsx to Supabase registrations table.")

if __name__ == "__main__":
    run()
