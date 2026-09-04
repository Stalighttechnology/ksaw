import openpyxl
import json
import urllib.request
import re

EXCEL_PATH = r"c:\Users\raghu\Desktop\ksaw\public\NEW XL. MALNAD.xlsx"
SUPABASE_URL = "https://wgtzcjsajncrvibtlhxv.supabase.co"
SUPABASE_KEY = "sb_publishable_ipyAcrU1KLwKoS20uWRBdA_iMVuIWtx"

headers = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
}

url = f"{SUPABASE_URL}/rest/v1/registrations?select=reference_number,first_name,last_name,phone,aadhaar_number"
req = urllib.request.Request(url, headers=headers)
with urllib.request.urlopen(req) as resp:
    db_rows = json.loads(resp.read().decode())

def clean_aadhaar(val):
    if not val:
        return ""
    return re.sub(r"\D", "", str(val))

db_map = {}
for r in db_rows:
    ref = r.get("reference_number") or ""
    digits = re.findall(r"\d+", ref)
    if digits and int(digits[0]) < 401:
        a = clean_aadhaar(r.get("aadhaar_number"))
        if a:
            db_map[a] = r

wb = openpyxl.load_workbook(EXCEL_PATH)
sheet = wb.active

seen_in_excel = {}
db_dups = []
intra_file_dups = []

for r in range(2, sheet.max_row + 1):
    name = str(sheet.cell(r, 3).value or "").strip()
    aadhaar_raw = sheet.cell(r, 15).value
    phone = str(sheet.cell(r, 14).value or "").replace(".0", "").strip()
    course = str(sheet.cell(r, 26).value or "").strip()
    a = clean_aadhaar(aadhaar_raw)
    
    if not a:
        continue
    
    if a in db_map:
        existing = db_map[a]
        db_dups.append({
            "excel_row": r,
            "name": name,
            "aadhaar": a,
            "phone": phone,
            "course": course,
            "db_ref": existing.get("reference_number"),
            "db_name": f"{existing.get('first_name')} {existing.get('last_name')}",
            "db_phone": existing.get("phone")
        })
    elif a in seen_in_excel:
        first_occurrence = seen_in_excel[a]
        intra_file_dups.append({
            "excel_row": r,
            "first_row": first_occurrence["excel_row"],
            "name": name,
            "aadhaar": a,
            "phone": phone,
            "course": course
        })
    else:
        seen_in_excel[a] = {
            "excel_row": r,
            "name": name,
            "phone": phone
        }

print("=== 1. DUPLICATES ALREADY IN DATABASE (7) ===")
for d in db_dups:
    print(f"Row {d['excel_row']}: {d['name']} | Aadhaar: {d['aadhaar']} | Phone: {d['phone']} --> Existing: [{d['db_ref']}: {d['db_name']}, Phone: {d['db_phone']}]")

print("\n=== 2. DUPLICATE ENTRIES REPEATED WITHIN EXCEL (8) ===")
for d in intra_file_dups:
    print(f"Row {d['excel_row']}: {d['name']} | Aadhaar: {d['aadhaar']} | Phone: {d['phone']} --> Duplicate of earlier Row {d['first_row']}")
